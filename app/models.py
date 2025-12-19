from datetime import datetime,timezone, timedelta
from sqlalchemy import Numeric
from email_service import email_service
from openpyxl import load_workbook

from extensions import db

from werkzeug.security import generate_password_hash, check_password_hash

# Инициализируем экземпляр SQLAlchemy

class User(db.Model):
    __tablename__ = 'users'

    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(512), nullable=False)  # для хранения хеша пароля
    role = db.Column(db.String(20), nullable=False, default='user')  # 'user', 'kgu', 'admin'
    confirmed = db.Column(db.Boolean, default=False)  # статус подтверждения СДО
    created_at = db.Column(db.DateTime, default=datetime.now(timezone.utc))

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def is_admin(self):
        return self.role == 'admin'

    def is_confirmed(self):
        return self.role == 'kgu'
    
    def __repr__(self):
        return f'<{self.role} {self.email}>'
    def to_dict(self):
        return {
            'id': self.id,
            'email': self.email,
            'role': self.role,
            'confirmed': self.confirmed,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }

class Notification(db.Model):
    __tablename__ = 'notifications'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    subject = db.Column(db.Text, default=id)
    message = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now(timezone.utc))
    status = db.Column(db.String(50), nullable=False, default='created')

    def send(self):
        email = User.query.get(self.user_id).email
        if email_service.send_email(email,self.subject,self.message):
            self.status = "sent"
        else:
            self.status = 'send error'
        db.session.commit()
        return self.status

    def __repr__(self):
        return f'<Notification {self.id}>'
    def to_dict(self):
        return {
            'id': self.id,
            'user_id': self.user_id,
            'subject': self.subject,
            'message': self.message,
            'status': self.status,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }

class Classroom(db.Model):
    __tablename__ = 'classrooms'

    id = db.Column(db.Integer, primary_key=True)
    eios_id =db.Column(db.Integer)
    number = db.Column(db.String(20))
    equipment = db.Column(db.Text)
    capacity = db.Column(db.Integer)
    description = db.Column(db.Text)
    price = db.Column(db.Integer)
    bookings = db.relationship('Booking', backref='classroom', lazy=True)

    def get_cost(self, hours):
        return self.price*hours

    def __repr__(self):
        return f'<Classroom {self.number}>'
    
    def to_dict(self, include_bookings=False):
        data = {
            'id': self.id,
            'eios_id': self.eios_id,
            'number': self.number,
            'equipment': self.equipment,
            'capacity': self.capacity,
            'description': self.description,
            'price': self.price
        }

        if include_bookings:
            data['bookings'] = [b.to_dict() for b in self.bookings]

        return data

class Booking(db.Model):
    __tablename__ = 'bookings'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classrooms.id'), nullable=False)
    date_start = db.Column(db.DateTime, nullable=False)
    date_end = db.Column(db.DateTime, nullable=False)

    status = db.Column(db.Boolean, default=None) # Null - на рассмотрении True/False - одобрено/отклонено
    description = db.Column(db.Text)

    created_at = db.Column(db.DateTime, default=datetime.now(timezone.utc))
    total_cost = db.Column(db.Numeric(10, 2),default=0.0)  # общая стоимость бронирования
    
    def get_duration(self):
        return self.date_end - self.date_start
    
    def __repr__(self):
        return f'<Бронирование {self.id}>'
    
    def to_dict(self, include_classroom=False):
        data = {
            'id': self.id,
            'user_id': self.user_id,
            'classroom_id': self.classroom_id,
            'date_start': self.date_start.isoformat() if self.date_start else None,
            'date_end': self.date_end.isoformat() if self.date_end else None,
            'status': self.status,
            'description': self.description,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'total_cost': float(self.total_cost) if self.total_cost is not None else 0.0,
            'duration_minutes': int(self.get_duration().total_seconds() // 60)
        }

        if include_classroom:
            data['classroom'] = self.classroom.to_dict()

        return data

def init_db():
    """Инициализация базы данных - создание всех таблиц"""
    from sqlalchemy import inspect

    # Создаем инспектор для проверки существования таблиц
    inspector = inspect(db.engine)
    existing_tables = inspector.get_table_names()

    # Проверяем, есть ли уже наши таблицы
    required_tables = ['users', 'classrooms', 'bookings', 'notifications']
    tables_exist = all(table in existing_tables for table in required_tables)

    if not tables_exist:
        db.create_all()
        print("✅ Все таблицы успешно созданы в PostgreSQL!")
    else:
        print("✅ Таблицы уже существуют в базе данных")



def parse_classrooms_xlsx(file):
    """
    file — это либо путь к файлу, либо FileStorage (из Flask request.files)
    """

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    # читаем заголовки
    headers = [
        cell.value.strip().lower()
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    classrooms = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        classroom = Classroom(
            eios_id=data.get("eios_id"),
            number=data.get("number"),
            equipment=data.get("equipment"),
            capacity=data.get("capacity"),
            description=data.get("description"),
            price=data.get("price"),
        )

        classrooms.append(classroom)
    return classrooms


def add_sample_data():
    try:
        db.session.query(User).delete()
        db.session.query(Notification).delete()
        db.session.query(Classroom).delete()
        db.session.query(Booking).delete()

        users = [
            User(email="someone@mail.com",password_hash=generate_password_hash("1234")),
            User(email="student@university.edu",confirmed=True,password_hash=generate_password_hash("1234")),
            User(role="moderator", email="moderator@university.edu",password_hash=generate_password_hash("1234")),
            User(role="admin", email="123@university.edu",password_hash=generate_password_hash("1234"))
        ]
        classrooms = parse_classrooms_xlsx("events/auds.xlsx")
    
        bookings = [
            Booking(
                classroom_id = 1,
                user_id = 1,
                date_start = datetime.now(timezone.utc) - timedelta(days=2),
                date_end = datetime.now(timezone.utc) - timedelta(days=2) + timedelta(hours=3),
                description = "Конференция",
                status=True,
                total_cost=4500.00
            ),
            Booking(
                classroom_id = 2,
                user_id = 2,
                date_start = datetime.now(timezone.utc) - timedelta(days=1),
                date_end = datetime.now(timezone.utc) - timedelta(days=1) + timedelta(hours=1.5),
                description = "мероприятие",
                status=True
            ),
        ]
        
        db.session.add_all(classrooms)
        db.session.add_all(users)
        db.session.add_all(bookings)
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Ошибка при добавлении тестовых данных: {e}")



WEB_ABLE_MODELS = {
    "user": User,
    "classroom": Classroom,
    "notification": Notification,
    "booking": Booking
}